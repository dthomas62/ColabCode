import io
import time
import random
import json
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import google.genai as genai
from google.cloud import storage, bigquery, secretmanager
from google.api_core import exceptions

# =====================================================================
# Configuration
# =====================================================================
PROJECT_ID = 'converged-brandradar-poc'
DATASET_ID = "brandradar"
REGION = 'europe-west2'
SECRET_ID = "GEMINI_API_KEY"

# Model and Parameters
MODELNAME = 'models/gemini-3.1-pro-preview'
BRAND = True
SOCIAL = True
CULTURE = True
REPORTNAME = 'Moretti 230226 '
DATASOURCE = 'gs://converged-brandradar/lookups/Moretti.csv'

# Templates
TEMPLATES = {
    'Brand': 'gs://converged-brandradar/templates/BrandRadar_Template.xlsx',
    'Social': 'gs://converged-brandradar/templates/SocialRadar_Template.xlsx',
    'Culture': 'gs://converged-brandradar/templates/CultureConverts_template.xlsx'
}

# =====================================================================
# Utilities
# =====================================================================
def access_secret_version(project_id: str, secret_id: str, version_id: str = "latest") -> str:
    """Accesses the payload for the given secret version."""
    client = secretmanager.SecretManagerServiceClient()
    name = f"projects/{project_id}/secrets/{secret_id}/versions/{version_id}"
    try:
        response = client.access_secret_version(request={"name": name})
        return response.payload.data.decode("UTF-8")
    except Exception as e:
        print(f"Error accessing secret: {e}")
        return None

def extract_json_robust(response_text):
    """Robust JSON extraction with multiple fallback strategies."""
    if not response_text:
        return None
        
    for pattern in [r'```json\s*(\{.*?\})\s*```', r'(\{.*\})']:
        match = re.search(pattern, response_text, re.DOTALL)
        if match:
            try:
                return json.loads(match.group(1))
            except json.JSONDecodeError:
                pass
                
    try:
        return json.loads(response_text)
    except json.JSONDecodeError:
        return None

def retry_with_backoff(retries=5, backoff_in_seconds=1):
    def decorator(f):
        def wrapper(*args, **kwargs):
            _retries, _backoff = retries, backoff_in_seconds
            while _retries > 1:
                try:
                    return f(*args, **kwargs)
                except exceptions.ResourceExhausted as e:
                    sleep = _backoff + random.uniform(0, 1)
                    print(f"API quota exceeded. Retrying in {sleep:.2f}s...")
                    time.sleep(sleep)
                    _backoff *= 2
                    _retries -= 1
                except Exception as e:
                    print(f"An unexpected error occurred: {e}. Retrying...")
                    time.sleep(_backoff)
                    _retries -= 1
            return f(*args, **kwargs)
        return wrapper
    return decorator

def get_question(task_name, brand, category, market):
    if task_name == 'Brand':
        return f"""Please provide a brief (approximately 200 words) summary of the {brand} brand's perception in the {category} category in the {market} market.

        For each of the following aspects, provide a score out of a 100 with a very short description of the score:
        - How often is it mentioned?
        - What is your overall impression of the brand?
        - Have people purchased the brand?
        - Is it thought of as a premium brand that people would pay a premium price for?
        - It is considered a sustainable brand?
        - do people trust it?
        - Would they recommend it to others?
        - What do you think of it's speed, dynamism and willingness to adapt to change?
        - Does the brand meet functional requirements?
        - Does the brand meet personal requirements?
        - Does the brand meet collective requirements?

        Could the results be returned in JSON format with fields: brand, market, summary, mentions, mentions_description, overallimp, overallimp_description, purchase, purchase_description, premium, premium_description, sustainable, sustainable_description, trust, trust_description, recommend, recommend_description, dynamic, dynamic_description, functional, functional_description, personal, personal_description, collective, collective_description, touchpoints_description, channel_description, journey_description, branded_description, sponsorship_description, cgc_description, brand_generated_description?"""
        
    elif task_name == 'Social':
        return f"""Please provide a brief (approximately 200 words) summary of the {brand} brand’s public sentiment and online conversation dynamics in the {category} category in the {market} market based on recent digital signals.

        For each of the following aspects, provide a score out of 100 and a short description explaining the score:
        - What is the general sentiment about the brand?
        - What emotions are most commonly associated with the brand?
        - Is the brand currently viewed more positively overall?
        - Is the brand currently viewed more negatively overall?
        - Are there noticeable sentiment shifts over the last 6 months?
        - Have there been sentiment spikes triggered by campaigns, news, or events?
        - What are the most discussed themes in online conversations about the brand?
        - What social, cultural, product or service topics are emerging?
        - Are there any polarizing or divisive conversations surrounding the brand?
        - Which platforms host the most conversations about the brand?
        - Where does the brand get the most engagement?
        - Is brand visibility driven more by owned channels or third-party/earned media?

        Please return results in **JSON format** using fields: "brand", "market", "summary", and "scores" object with sentiment, emotion, positivity, sentiment_shift, sentiment_spike, themes, emerging_topics, debates, platforms, engagement, media_type."""
        
    elif task_name == 'Culture':
        return f"""Please analyze the {brand} brand's perception in the {category} category in the {market} market.
        I am interested in understanding 4 areas, buzz, belonging, belief and behaviour. Could provide a value between 1 and 100 for the following questions and store them in JSON format with fields called Buzz1 etc?
        Buzz1 to Buzz5, Belong1 to Belong5, Belief1 to Belief5, Behave1 to Behave5.
        
        CRITICAL: Return ONLY a valid JSON object with NO markdown formatting, NO additional text."""

# =====================================================================
# State and Schema Definitions
# =====================================================================
ANALYSIS_CONFIGS = {
    'Brand': {
        'run': BRAND,
        'table_id': 'BrandResults',
        'sheet': 'BrandRadar Results',
        'template': TEMPLATES['Brand'],
        'int_cols': ["mentions", "overallimp", "purchase", "premium", "sustainable", 
                     "trust", "recommend", "dynamic", "functional", "personal", "collective"],
        'str_cols': ["touchpoints", "channel", "journey", "branded", "sponsorship", "cgc", "brand_generated"]
    },
    'Social': {
        'run': SOCIAL,
        'table_id': 'SocialResults',
        'sheet': 'SocialRadar Results',
        'template': TEMPLATES['Social'],
        'int_cols': ["sentiment", "emotion", "positivity", "shift", "spike", "themes", 
                     "emerging_topics", "debates", "platforms", "engagement", "media_type"],
        'str_cols': []
    },
    'Culture': {
        'run': CULTURE,
        'table_id': 'CultureConverts',
        'sheet': 'Cultural Converts',
        'template': TEMPLATES['Culture'],
        'int_cols': [f"Buzz{i}" for i in range(1, 6)] + [f"Belong{i}" for i in range(1, 6)] + \
                    [f"Belief{i}" for i in range(1, 6)] + [f"Behave{i}" for i in range(1, 6)],
        'str_cols': []
    }
}

# Attach full columns list for dataframe initialization mapping
for t, c in ANALYSIS_CONFIGS.items():
    if t == 'Culture':
        c['columns'] = ["Brand", "Category", "Market"] + c['int_cols'] + ["date", "projectname"]
    else:
        c['columns'] = ["Brand", "Category", "Market", "Summary"] + \
                       c['int_cols'] + \
                       [f"{col}_description" for col in c['int_cols']] + \
                       [f"{col}_description" for col in c['str_cols']] + \
                       ["date"]

def initialize_dataframe(task_name):
    """Dynamically set up DataFrame based on schema."""
    cols = ANALYSIS_CONFIGS[task_name]['columns']
    df = pd.DataFrame(columns=cols)
    for col in ANALYSIS_CONFIGS[task_name]['int_cols']:
        df[col] = pd.Series(dtype='int64')
    df['date'] = pd.Series(dtype='datetime64[ns]')
    return df

# =====================================================================
# API Logic
# =====================================================================
@retry_with_backoff(retries=3, backoff_in_seconds=2)
def run_query(client, question):
    try:
        response = client.models.generate_content(model=MODELNAME, contents=question)
        if not response.candidates:
            print(f"Request blocked. Feedback: {response.prompt_feedback}")
            return None
        return response.text
    except ValueError:
        print("Response empty or blocked.")
        return None

def extract_values_to_row(task_name, json_data, row_template):
    """Parses logic correctly depending on specific task expectations."""
    if not json_data:
        return row_template
        
    config = ANALYSIS_CONFIGS[task_name]
    new_row = row_template.copy()

    # Generic mappings
    new_row['Summary'] = json_data.get('summary', '')

    if task_name == 'Social':
        scores = json_data.get('scores', {})
        for measure in config['int_cols']:
            # Adjust the key naming difference specific for social
            key = "sentiment_shift" if measure == "shift" else "sentiment_spike" if measure == "spike" else measure
            if key in scores:
                new_row[measure] = scores[key].get('score', 0)
                new_row[f"{measure}_description"] = scores[key].get('description', '')
    else:
        for k, v in json_data.items():
            if k in new_row or k not in config['columns']: continue
            new_row[k] = v

    return new_row

def process_record(client, row, task_name):
    """Processes a single brand query."""
    question = get_question(task_name, row['brand'], row['category'], row['market'])
    response_text = run_query(client, question)
    
    brand_json = extract_json_robust(response_text) if response_text else None
    
    # Init base row structure
    row_template = {
        "Brand": brand_json.get('brand', row['brand']) if brand_json else row['brand'],
        "Category": row['category'],
        "Market": row['market'],
        "date": datetime.now().strftime("%Y/%m/%d")
    }
    if task_name == 'Culture':
        row_template["projectname"] = REPORTNAME

    row_data = extract_values_to_row(task_name, brand_json, row_template)
        
    return row_data, question, response_text, bool(brand_json)

def execute_task(task_name, df_input, client, storage_client):
    """Executes the full pipeline for a specific reporting mode."""
    print(f"\n{'='*40}\nProcessing Task: {task_name}\n{'='*40}")
    
    df = initialize_dataframe(task_name)
    all_responses = []
    
    for idx, row in df_input.iterrows():
        print(f"[{idx+1}/{len(df_input)}] Processing {row['brand']} ({row['market']})")
        row_data, question, resp_text, success = process_record(client, row, task_name)
        
        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)
        
        all_responses.append({
            "timestamp": datetime.now().isoformat(),
            "brand": row['brand'],
            "market": row['market'],
            "input_question": question,
            "gemini_response": resp_text,
            "success": success
        })
        time.sleep(1) # Rate limiter
        
    df['projectname'] = REPORTNAME
        
    # Exporters
    bucket = storage_client.bucket("converged-brandradar")
    
    # 1. JSON
    blob_name = f"results/{task_name} {datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    bucket.blob(blob_name).upload_from_string(json.dumps(all_responses, indent=4), content_type='application/json')
    print(f"JSON Exported: gs://converged-brandradar/{blob_name}")

    # 2. XLS
    config = ANALYSIS_CONFIGS[task_name]
    xls_path = f"gs://converged-brandradar/results/{task_name} {REPORTNAME.strip()} {datetime.now().strftime('%Y%m%d')}.xlsx"
    
    template_bucket = storage_client.bucket(config['template'].split('/')[2])
    template_blob = template_bucket.blob('/'.join(config['template'].split('/')[3:]))
    template_buffer = io.BytesIO()
    template_blob.download_to_file(template_buffer)
    template_buffer.seek(0)
    
    book = load_workbook(template_buffer)
    book['Key'].cell(row=3, column=1, value=REPORTNAME)
    sheet = book[config['sheet']]
    
    for r_idx, out_row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(out_row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
            
    out_buffer = io.BytesIO()
    book.save(out_buffer)
    out_buffer.seek(0)
    
    out_bucket = storage_client.bucket(xls_path.split('/')[2])
    out_bucket.blob('/'.join(xls_path.split('/')[3:])).upload_from_file(out_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    print(f"Excel Exported: {xls_path}")

    # 3. BigQuery
    table_ref = f"{PROJECT_ID}.{DATASET_ID}.{config['table_id']}"
    schema = []
    for col in df.columns:
        dtype = df[col].dtype
        bq_type = "INTEGER" if dtype == 'int64' else "FLOAT" if dtype == 'float64' else "TIMESTAMP" if dtype == 'datetime64[ns]' else "STRING"
        schema.append(bigquery.SchemaField(col, bq_type))
        
    bq_client = bigquery.Client(project=PROJECT_ID)
    try:
        bq_client.get_table(table_ref)
        job_config = bigquery.LoadJobConfig(write_disposition=bigquery.WriteDisposition.WRITE_APPEND)
    except Exception:
        job_config = bigquery.LoadJobConfig(write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, schema=schema)
        
    bq_client.load_table_from_dataframe(df, table_ref, job_config=job_config).result()
    print(f"BigQuery Updated: {config['table_id']} ({len(df)} rows)")

# =====================================================================
# Entrypoint
# =====================================================================
def main():
    start_time = time.time()
    
    api_key = access_secret_version(PROJECT_ID, SECRET_ID)
    if not api_key:
        print("Failed to retrieve API key!")
        return
        
    client = genai.Client(api_key=api_key)
    storage_client = storage.Client()
    
    # Load and process Data
    dfbrands = pd.read_csv(DATASOURCE).head(1) # Drop head() modifier when ready
    cats = dfbrands.groupby(['market', 'category']).size().reset_index()
    cats['brand'] = 'All Category'
    dfbrands = pd.concat([dfbrands, cats], ignore_index=True)

    for task_name, config in ANALYSIS_CONFIGS.items():
        if config['run']:
            execute_task(task_name, dfbrands, client, storage_client)
            
    print(f"\nTotal Run time: {time.time() - start_time:.2f} seconds")

if __name__ == "__main__":
    main()
